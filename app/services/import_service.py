from __future__ import annotations

import csv
import re
from dataclasses import dataclass
from datetime import UTC, datetime, time
from enum import Enum, StrEnum
from io import BytesIO, StringIO
from pathlib import Path
from typing import Any, TypeVar

from openpyxl import load_workbook
from sqlalchemy import select

from app.database.models.bush import Bush
from app.database.models.cluster import Cluster
from app.database.models.enums import (
    AuditAction,
    EntityType,
)
from app.database.models.store import Store
from app.database.models.user import User
from app.repositories import (
    AuditContext,
    Repositories,
)
from app.services.access import AccessService
from app.services.file_service import (
    DownloadedFile,
)
from app.services.store_service import (
    StoreService,
)


EnumType = TypeVar(
    "EnumType",
    bound=Enum,
)


class ImportFileFormat(StrEnum):
    """
    Формат файлу імпорту.
    """

    XLSX = "xlsx"
    XLS = "xls"
    CSV = "csv"


class ImportRowStatus(StrEnum):
    """
    Що буде зроблено з рядком.
    """

    CREATE = "create"
    UPDATE = "update"
    UNCHANGED = "unchanged"

    IGNORED = "ignored"
    INVALID = "invalid"


class ImportIssueLevel(StrEnum):
    """
    Рівень проблеми.
    """

    WARNING = "warning"
    ERROR = "error"


@dataclass(slots=True, frozen=True)
class ImportIssue:
    """
    Помилка або попередження імпорту.
    """

    row_number: int | None

    field: str | None

    level: ImportIssueLevel

    message: str


@dataclass(slots=True, frozen=True)
class StoreImportRow:
    """
    Один нормалізований рядок імпорту.
    """

    row_number: int

    status: ImportRowStatus

    store_number: int | None
    code: str | None

    name: str | None
    city: str | None
    address: str | None

    bush_id: int | None
    cluster_id: int | None

    is_active: bool | None

    existing_store_id: int | None

    issues: tuple[
        ImportIssue,
        ...,
    ]

    raw_values: dict[
        str,
        Any,
    ]

    @property
    def is_valid(self) -> bool:
        return self.status not in {
            ImportRowStatus.INVALID,
        }

    @property
    def is_actionable(self) -> bool:
        return self.status in {
            ImportRowStatus.CREATE,
            ImportRowStatus.UPDATE,
        }

    @property
    def has_errors(self) -> bool:
        return any(
            issue.level
            == ImportIssueLevel.ERROR
            for issue in self.issues
        )


@dataclass(slots=True, frozen=True)
class ImportPreview:
    """
    Preview перед записом у PostgreSQL.
    """

    file_name: str

    file_format: ImportFileFormat

    sheet_name: str | None

    header_row_number: int

    total_rows: int

    create_count: int
    update_count: int
    unchanged_count: int
    ignored_count: int
    invalid_count: int

    rows: tuple[
        StoreImportRow,
        ...,
    ]

    issues: tuple[
        ImportIssue,
        ...,
    ]

    @property
    def actionable_count(self) -> int:
        return (
            self.create_count
            + self.update_count
        )

    @property
    def valid_count(self) -> int:
        return (
            self.total_rows
            - self.invalid_count
        )

    @property
    def can_import(self) -> bool:
        return (
            self.actionable_count > 0
            and self.invalid_count == 0
        )


@dataclass(slots=True, frozen=True)
class ImportApplyItemResult:
    """
    Результат запису одного рядка.
    """

    row_number: int

    code: str | None
    store_number: int | None

    requested_status: ImportRowStatus

    success: bool

    store_id: int | None

    was_created: bool
    was_updated: bool

    error: str | None


@dataclass(slots=True, frozen=True)
class ImportApplyResult:
    """
    Результат фактичного імпорту.
    """

    file_name: str

    total_preview_rows: int

    attempted_count: int

    success_count: int
    failed_count: int

    created_count: int
    updated_count: int

    unchanged_count: int
    ignored_count: int
    invalid_count: int

    items: tuple[
        ImportApplyItemResult,
        ...,
    ]

    imported_at: datetime
    imported_by_id: int


@dataclass(slots=True, frozen=True)
class ParsedTable:
    """
    Сирі дані таблиці.
    """

    file_format: ImportFileFormat

    sheet_name: str | None

    header_row_number: int

    headers: tuple[
        str,
        ...,
    ]

    rows: tuple[
        tuple[int, tuple[Any, ...]],
        ...,
    ]


class ImportService:
    """
    Сервіс імпорту торгових точок.

    Підтримує:

        XLSX
        XLS
        CSV

    Основний процес:

        файл
          ↓
        preview
          ↓
        перевірка
          ↓
        CREATE / UPDATE / IGNORE
          ↓
        підтвердження адміном
          ↓
        PostgreSQL

    Важливо:

    Київ автоматично виключається
    з імпорту.

    Фізичне видалення ТТ
    через імпорт не виконується.
    """

    MAX_ROWS = 5000

    HEADER_SCAN_ROWS = 25

    KYIV_NAMES = frozenset(
        {
            "київ",
            "киев",
            "kyiv",
            "kiev",
        }
    )

    HEADER_ALIASES: dict[
        str,
        frozenset[str],
    ] = {
        "store_number": frozenset(
            {
                "номер",
                "№",
                "№ тт",
                "номер тт",
                "номер магазину",
                "номер магазина",
                "store number",
                "store_number",
                "sb",
                "id тт",
            }
        ),

        "code": frozenset(
            {
                "код",
                "код тт",
                "код магазину",
                "код магазина",
                "store code",
                "store_code",
                "code",
                "тт",
            }
        ),

        "name": frozenset(
            {
                "назва",
                "назва тт",
                "назва магазину",
                "название",
                "name",
                "store name",
            }
        ),

        "city": frozenset(
            {
                "місто",
                "город",
                "місто/село",
                "населений пункт",
                "населенный пункт",
                "city",
            }
        ),

        "address": frozenset(
            {
                "адреса",
                "адрес",
                "адреса тт",
                "адреса магазину",
                "address",
            }
        ),

        "bush": frozenset(
            {
                "кущ",
                "куст",
                "кущ id",
                "кущ_id",
                "bush",
                "bush id",
            }
        ),

        "cluster": frozenset(
            {
                "кластер",
                "cluster",
                "час відкриття",
                "час открытия",
                "графік відкриття",
                "opening",
                "opening time",
            }
        ),

        "is_active": frozenset(
            {
                "активний",
                "активна",
                "active",
                "is active",
                "is_active",
                "статус",
                "status",
            }
        ),
    }

    def __init__(
        self,
        repositories: Repositories,
        *,
        access_service: AccessService | None = None,
        store_service: StoreService | None = None,
    ) -> None:
        self.repositories = repositories
        self.session = repositories.session

        self.access = (
            access_service
            or AccessService(repositories)
        )

        self.stores = (
            store_service
            or StoreService(
                repositories,
                access_service=self.access,
            )
        )

        self.header_lookup = (
            self.build_header_lookup()
        )

    # ==========================================
    # PREVIEW DOWNLOADED FILE
    # ==========================================

    async def preview_downloaded(
        self,
        *,
        actor: User,
        downloaded: DownloadedFile,
        sheet_name: str | None = None,
        ignore_kyiv: bool = True,
    ) -> ImportPreview:
        """
        Preview Telegram-файлу.
        """

        return await self.preview_bytes(
            actor=actor,
            content=downloaded.content,
            file_name=downloaded.file_name,
            sheet_name=sheet_name,
            ignore_kyiv=ignore_kyiv,
        )

    # ==========================================
    # PREVIEW BYTES
    # ==========================================

    async def preview_bytes(
        self,
        *,
        actor: User,
        content: bytes,
        file_name: str,
        sheet_name: str | None = None,
        ignore_kyiv: bool = True,
    ) -> ImportPreview:
        """
        Розбирає файл без запису в БД.
        """

        self.access.require_network_management(
            actor
        )

        if not content:
            raise ValueError(
                "Файл порожній."
            )

        parsed = self.parse_table(
            content=content,
            file_name=file_name,
            sheet_name=sheet_name,
        )

        existing_stores = (
            await self.load_existing_stores()
        )

        bushes = await self.load_bushes()
        clusters = await self.load_clusters()

        store_code_map = {
            self.store_code(store).lower():
            store
            for store in existing_stores
        }

        store_number_map: dict[
            int,
            Store,
        ] = {}

        for store in existing_stores:
            number = self.store_number(
                store
            )

            if number is not None:
                store_number_map[
                    number
                ] = store

        bush_maps = self.build_bush_maps(
            bushes
        )

        cluster_maps = (
            self.build_cluster_maps(
                clusters
            )
        )

        seen_codes: dict[
            str,
            int,
        ] = {}

        seen_numbers: dict[
            int,
            int,
        ] = {}

        preview_rows: list[
            StoreImportRow
        ] = []

        global_issues: list[
            ImportIssue
        ] = []

        header_mapping = (
            self.resolve_header_mapping(
                parsed.headers
            )
        )

        if (
            "code" not in header_mapping
            and "store_number"
            not in header_mapping
        ):
            raise ValueError(
                "У файлі повинна бути колонка "
                "«Код» або «Номер ТТ»."
            )

        for (
            row_number,
            raw_row,
        ) in parsed.rows:
            if self.row_is_empty(
                raw_row
            ):
                continue

            if (
                len(preview_rows)
                >= self.MAX_ROWS
            ):
                global_issues.append(
                    ImportIssue(
                        row_number=None,
                        field=None,
                        level=(
                            ImportIssueLevel.ERROR
                        ),
                        message=(
                            "Перевищено максимальну "
                            f"кількість рядків: "
                            f"{self.MAX_ROWS}."
                        ),
                    )
                )

                break

            raw_values = (
                self.map_row_values(
                    raw_row=raw_row,
                    header_mapping=(
                        header_mapping
                    ),
                )
            )

            row = self.parse_import_row(
                row_number=row_number,
                raw_values=raw_values,

                store_code_map=(
                    store_code_map
                ),
                store_number_map=(
                    store_number_map
                ),

                bush_maps=bush_maps,

                cluster_maps=(
                    cluster_maps
                ),

                seen_codes=seen_codes,
                seen_numbers=seen_numbers,

                ignore_kyiv=ignore_kyiv,
            )

            preview_rows.append(
                row
            )

        all_issues = list(
            global_issues
        )

        for row in preview_rows:
            all_issues.extend(
                row.issues
            )

        return ImportPreview(
            file_name=file_name,

            file_format=(
                parsed.file_format
            ),

            sheet_name=(
                parsed.sheet_name
            ),

            header_row_number=(
                parsed.header_row_number
            ),

            total_rows=len(
                preview_rows
            ),

            create_count=sum(
                row.status
                == ImportRowStatus.CREATE
                for row in preview_rows
            ),

            update_count=sum(
                row.status
                == ImportRowStatus.UPDATE
                for row in preview_rows
            ),

            unchanged_count=sum(
                row.status
                == ImportRowStatus.UNCHANGED
                for row in preview_rows
            ),

            ignored_count=sum(
                row.status
                == ImportRowStatus.IGNORED
                for row in preview_rows
            ),

            invalid_count=sum(
                row.status
                == ImportRowStatus.INVALID
                for row in preview_rows
            ),

            rows=tuple(
                preview_rows
            ),

            issues=tuple(
                all_issues
            ),
        )

    # ==========================================
    # PARSE ONE ROW
    # ==========================================

    def parse_import_row(
        self,
        *,
        row_number: int,

        raw_values: dict[str, Any],

        store_code_map: dict[str, Store],
        store_number_map: dict[int, Store],

        bush_maps: dict[str, dict[Any, Bush]],
        cluster_maps: dict[
            str,
            dict[Any, Cluster],
        ],

        seen_codes: dict[str, int],
        seen_numbers: dict[int, int],

        ignore_kyiv: bool,
    ) -> StoreImportRow:
        """
        Нормалізує один рядок.
        """

        issues: list[
            ImportIssue
        ] = []

        # --------------------------------------
        # NUMBER
        # --------------------------------------

        store_number = (
            self.parse_store_number(
                raw_values.get(
                    "store_number"
                )
            )
        )

        # --------------------------------------
        # CODE
        # --------------------------------------

        code_raw = raw_values.get(
            "code"
        )

        code: str | None = None

        if self.has_value(code_raw):
            try:
                code = (
                    StoreService.normalize_code(
                        str(code_raw)
                    )
                )

            except ValueError as error:
                issues.append(
                    ImportIssue(
                        row_number=row_number,
                        field="code",
                        level=(
                            ImportIssueLevel.ERROR
                        ),
                        message=str(error),
                    )
                )

        # --------------------------------------
        # INFER NUMBER FROM CODE
        # --------------------------------------

        if (
            store_number is None
            and code is not None
        ):
            store_number = (
                self.number_from_code(
                    code
                )
            )

        # --------------------------------------
        # INFER CODE FROM NUMBER
        # --------------------------------------

        if (
            code is None
            and store_number is not None
        ):
            code = (
                f"SB-{store_number}"
            )

        if (
            code is None
            or store_number is None
        ):
            issues.append(
                ImportIssue(
                    row_number=row_number,
                    field=(
                        "code/store_number"
                    ),
                    level=(
                        ImportIssueLevel.ERROR
                    ),
                    message=(
                        "Не вдалося визначити "
                        "номер торгової точки."
                    ),
                )
            )

        # --------------------------------------
        # DUPLICATE IN FILE
        # --------------------------------------

        if code is not None:
            code_key = code.lower()

            duplicate_row = (
                seen_codes.get(
                    code_key
                )
            )

            if duplicate_row is not None:
                issues.append(
                    ImportIssue(
                        row_number=row_number,
                        field="code",
                        level=(
                            ImportIssueLevel.ERROR
                        ),
                        message=(
                            f"Дубль {code} у файлі. "
                            "Перший рядок: "
                            f"{duplicate_row}."
                        ),
                    )
                )

            else:
                seen_codes[
                    code_key
                ] = row_number

        if store_number is not None:
            duplicate_row = (
                seen_numbers.get(
                    store_number
                )
            )

            if duplicate_row is not None:
                issues.append(
                    ImportIssue(
                        row_number=row_number,
                        field="store_number",
                        level=(
                            ImportIssueLevel.ERROR
                        ),
                        message=(
                            "Дубль номера "
                            f"{store_number}. "
                            "Перший рядок: "
                            f"{duplicate_row}."
                        ),
                    )
                )

            else:
                seen_numbers[
                    store_number
                ] = row_number

        # --------------------------------------
        # EXISTING STORE
        # --------------------------------------

        existing_by_code = (
            store_code_map.get(
                code.lower()
            )
            if code
            else None
        )

        existing_by_number = (
            store_number_map.get(
                store_number
            )
            if store_number is not None
            else None
        )

        if (
            existing_by_code is not None
            and existing_by_number is not None
            and existing_by_code.id
            != existing_by_number.id
        ):
            issues.append(
                ImportIssue(
                    row_number=row_number,
                    field=(
                        "code/store_number"
                    ),
                    level=(
                        ImportIssueLevel.ERROR
                    ),
                    message=(
                        "Код і номер відповідають "
                        "різним ТТ у базі."
                    ),
                )
            )

        existing = (
            existing_by_code
            or existing_by_number
        )

        # --------------------------------------
        # CITY
        # --------------------------------------

        city_from_file = (
            self.normalize_optional_text(
                raw_values.get(
                    "city"
                )
            )
        )

        existing_city = (
            self.get_text_attribute(
                existing,
                "city",
            )
            if existing is not None
            else None
        )

        city = (
            city_from_file
            or existing_city
        )

        if (
            existing is None
            and not city
        ):
            issues.append(
                ImportIssue(
                    row_number=row_number,
                    field="city",
                    level=(
                        ImportIssueLevel.ERROR
                    ),
                    message=(
                        "Для нової ТТ "
                        "не вказано місто."
                    ),
                )
            )

        # --------------------------------------
        # IGNORE KYIV
        # --------------------------------------

        if (
            ignore_kyiv
            and city
            and self.is_kyiv(
                city
            )
        ):
            return StoreImportRow(
                row_number=row_number,

                status=(
                    ImportRowStatus.IGNORED
                ),

                store_number=store_number,
                code=code,

                name=(
                    self.normalize_optional_text(
                        raw_values.get(
                            "name"
                        )
                    )
                ),

                city=city,

                address=(
                    self.normalize_optional_text(
                        raw_values.get(
                            "address"
                        )
                    )
                ),

                bush_id=None,
                cluster_id=None,

                is_active=None,

                existing_store_id=(
                    existing.id
                    if existing
                    else None
                ),

                issues=(
                    ImportIssue(
                        row_number=row_number,
                        field="city",
                        level=(
                            ImportIssueLevel.WARNING
                        ),
                        message=(
                            "ТТ Києва автоматично "
                            "виключена з імпорту."
                        ),
                    ),
                ),

                raw_values=raw_values,
            )

        # --------------------------------------
        # NAME
        # --------------------------------------

        name_from_file = (
            self.normalize_optional_text(
                raw_values.get(
                    "name"
                )
            )
        )

        existing_name = (
            self.store_name(
                existing
            )
            if existing is not None
            else None
        )

        name = (
            name_from_file
            or existing_name
        )

        if (
            name is None
            and code
        ):
            name = (
                f"{code}"
                + (
                    f" {city}"
                    if city
                    else ""
                )
            )

        # --------------------------------------
        # ADDRESS
        # --------------------------------------

        address_from_file = (
            self.normalize_optional_text(
                raw_values.get(
                    "address"
                )
            )
        )

        existing_address = (
            self.get_text_attribute(
                existing,
                "address",
            )
            if existing is not None
            else None
        )

        address = (
            address_from_file
            if address_from_file is not None
            else existing_address
        )

        # --------------------------------------
        # BUSH
        # --------------------------------------

        bush_value = raw_values.get(
            "bush"
        )

        if self.has_value(
            bush_value
        ):
            bush_id = (
                self.resolve_bush(
                    bush_value,
                    bush_maps,
                )
            )

            if bush_id is None:
                issues.append(
                    ImportIssue(
                        row_number=row_number,
                        field="bush",
                        level=(
                            ImportIssueLevel.ERROR
                        ),
                        message=(
                            "Не знайдено кущ: "
                            f"{bush_value}."
                        ),
                    )
                )

        else:
            bush_id = (
                getattr(
                    existing,
                    "bush_id",
                    None,
                )
                if existing
                else None
            )

        # --------------------------------------
        # CLUSTER
        # --------------------------------------

        cluster_value = raw_values.get(
            "cluster"
        )

        if self.has_value(
            cluster_value
        ):
            cluster_id = (
                self.resolve_cluster(
                    cluster_value,
                    cluster_maps,
                )
            )

            if cluster_id is None:
                issues.append(
                    ImportIssue(
                        row_number=row_number,
                        field="cluster",
                        level=(
                            ImportIssueLevel.ERROR
                        ),
                        message=(
                            "Не знайдено кластер: "
                            f"{cluster_value}."
                        ),
                    )
                )

        else:
            cluster_id = (
                getattr(
                    existing,
                    "cluster_id",
                    None,
                )
                if existing
                else None
            )

        # --------------------------------------
        # ACTIVE
        # --------------------------------------

        active_value = raw_values.get(
            "is_active"
        )

        if self.has_value(
            active_value
        ):
            try:
                is_active = (
                    self.parse_bool(
                        active_value
                    )
                )

            except ValueError as error:
                issues.append(
                    ImportIssue(
                        row_number=row_number,
                        field="is_active",
                        level=(
                            ImportIssueLevel.ERROR
                        ),
                        message=str(error),
                    )
                )

                is_active = None

        else:
            is_active = (
                bool(
                    getattr(
                        existing,
                        "is_active",
                        True,
                    )
                )
                if existing is not None
                else True
            )

        # --------------------------------------
        # INVALID?
        # --------------------------------------

        has_errors = any(
            issue.level
            == ImportIssueLevel.ERROR
            for issue in issues
        )

        if has_errors:
            status = (
                ImportRowStatus.INVALID
            )

        elif existing is None:
            status = (
                ImportRowStatus.CREATE
            )

        else:
            status = (
                ImportRowStatus.UPDATE
                if self.store_needs_update(
                    store=existing,

                    store_number=(
                        store_number
                    ),
                    code=code,
                    name=name,
                    city=city,
                    address=address,

                    bush_id=bush_id,
                    cluster_id=(
                        cluster_id
                    ),

                    is_active=(
                        is_active
                    ),
                )
                else (
                    ImportRowStatus.UNCHANGED
                )
            )

        return StoreImportRow(
            row_number=row_number,

            status=status,

            store_number=(
                store_number
            ),

            code=code,

            name=name,
            city=city,
            address=address,

            bush_id=bush_id,
            cluster_id=cluster_id,

            is_active=is_active,

            existing_store_id=(
                existing.id
                if existing
                else None
            ),

            issues=tuple(
                issues
            ),

            raw_values=raw_values,
        )

    # ==========================================
    # APPLY PREVIEW
    # ==========================================

    async def apply_preview(
        self,
        *,
        actor: User,
        preview: ImportPreview,
        allow_partial: bool = False,
        reason: str = (
            "Імпорт торгових точок"
        ),
    ) -> ImportApplyResult:
        """
        Фактично записує preview у PostgreSQL.
        """

        self.access.require_network_management(
            actor
        )

        normalized_reason = (
            self.normalize_required_text(
                reason,
                field_name="Причина",
                max_length=2000,
            )
        )

        if (
            preview.invalid_count > 0
            and not allow_partial
        ):
            raise ValueError(
                "У preview є помилки. "
                "Виправте файл або використайте "
                "частковий імпорт."
            )

        imported_at = datetime.now(
            UTC
        )

        results: list[
            ImportApplyItemResult
        ] = []

        actionable_rows = [
            row
            for row in preview.rows
            if row.is_actionable
        ]

        for row in actionable_rows:
            try:
                async with self.session.begin_nested():
                    item = (
                        await self.apply_row(
                            actor=actor,
                            row=row,
                            reason=(
                                normalized_reason
                            ),
                            changed_at=(
                                imported_at
                            ),
                        )
                    )

                results.append(
                    item
                )

            except Exception as error:
                results.append(
                    ImportApplyItemResult(
                        row_number=(
                            row.row_number
                        ),
                        code=row.code,
                        store_number=(
                            row.store_number
                        ),
                        requested_status=(
                            row.status
                        ),
                        success=False,
                        store_id=(
                            row.existing_store_id
                        ),
                        was_created=False,
                        was_updated=False,
                        error=str(error),
                    )
                )

                if not allow_partial:
                    raise RuntimeError(
                        "Імпорт зупинено "
                        f"на рядку {row.row_number}: "
                        f"{error}"
                    ) from error

        result = ImportApplyResult(
            file_name=preview.file_name,

            total_preview_rows=(
                preview.total_rows
            ),

            attempted_count=len(
                actionable_rows
            ),

            success_count=sum(
                item.success
                for item in results
            ),

            failed_count=sum(
                not item.success
                for item in results
            ),

            created_count=sum(
                item.was_created
                for item in results
            ),

            updated_count=sum(
                item.was_updated
                for item in results
            ),

            unchanged_count=(
                preview.unchanged_count
            ),

            ignored_count=(
                preview.ignored_count
            ),

            invalid_count=(
                preview.invalid_count
            ),

            items=tuple(
                results
            ),

            imported_at=imported_at,
            imported_by_id=actor.id,
        )

        await self.log_import(
            actor=actor,
            preview=preview,
            result=result,
            reason=normalized_reason,
        )

        return result

    # ==========================================
    # APPLY ONE ROW
    # ==========================================

    async def apply_row(
        self,
        *,
        actor: User,
        row: StoreImportRow,
        reason: str,
        changed_at: datetime,
    ) -> ImportApplyItemResult:
        """
        Записує один рядок.
        """

        if not row.is_actionable:
            raise ValueError(
                "Рядок не потребує імпорту."
            )

        if (
            row.store_number is None
            or row.code is None
            or row.city is None
        ):
            raise ValueError(
                "У рядку відсутні "
                "обов’язкові дані."
            )

        # --------------------------------------
        # CREATE
        # --------------------------------------

        if (
            row.status
            == ImportRowStatus.CREATE
        ):
            create_result = (
                await self.stores.create_store(
                    actor=actor,

                    store_number=(
                        row.store_number
                    ),

                    code=row.code,

                    name=row.name,

                    city=row.city,

                    address=row.address,

                    bush_id=row.bush_id,

                    cluster_id=(
                        row.cluster_id
                    ),

                    is_active=(
                        True
                        if row.is_active is None
                        else row.is_active
                    ),

                    reason=reason,

                    created_at=(
                        changed_at
                    ),
                )
            )

            return ImportApplyItemResult(
                row_number=row.row_number,

                code=row.code,

                store_number=(
                    row.store_number
                ),

                requested_status=(
                    row.status
                ),

                success=True,

                store_id=(
                    create_result.store.id
                ),

                was_created=True,
                was_updated=False,

                error=None,
            )

        # --------------------------------------
        # UPDATE
        # --------------------------------------

        if row.existing_store_id is None:
            raise ValueError(
                "Не визначено існуючу ТТ."
            )

        store = (
            await self.stores.get_store_or_raise(
                row.existing_store_id,
                for_update=True,
                include_inactive=True,
            )
        )

        await self.stores.update_store(
            actor=actor,

            store_id=store.id,

            name=row.name,

            city=row.city,

            address=row.address,

            code=row.code,

            store_number=(
                row.store_number
            ),

            reason=reason,

            updated_at=changed_at,
        )

        current_bush_id = getattr(
            store,
            "bush_id",
            None,
        )

        if (
            current_bush_id
            != row.bush_id
        ):
            await self.stores.change_bush(
                actor=actor,

                store_id=store.id,

                bush_id=row.bush_id,

                reason=reason,

                changed_at=changed_at,
            )

        current_cluster_id = getattr(
            store,
            "cluster_id",
            None,
        )

        if (
            current_cluster_id
            != row.cluster_id
        ):
            await self.stores.change_cluster(
                actor=actor,

                store_id=store.id,

                cluster_id=(
                    row.cluster_id
                ),

                reason=reason,

                changed_at=changed_at,
            )

        if row.is_active is not None:
            current_active = bool(
                getattr(
                    store,
                    "is_active",
                    True,
                )
            )

            if (
                current_active
                != row.is_active
            ):
                await self.stores.set_active_state(
                    actor=actor,

                    store_id=store.id,

                    is_active=(
                        row.is_active
                    ),

                    reason=reason,

                    deactivate_bindings=False,

                    changed_at=changed_at,
                )

        return ImportApplyItemResult(
            row_number=row.row_number,

            code=row.code,

            store_number=(
                row.store_number
            ),

            requested_status=row.status,

            success=True,

            store_id=store.id,

            was_created=False,
            was_updated=True,

            error=None,
        )

    # ==========================================
    # PARSE TABLE
    # ==========================================

    def parse_table(
        self,
        *,
        content: bytes,
        file_name: str,
        sheet_name: str | None,
    ) -> ParsedTable:
        """
        Визначає формат і читає таблицю.
        """

        file_format = (
            self.detect_file_format(
                file_name
            )
        )

        if (
            file_format
            == ImportFileFormat.XLSX
        ):
            return self.parse_xlsx(
                content=content,
                sheet_name=sheet_name,
            )

        if (
            file_format
            == ImportFileFormat.CSV
        ):
            return self.parse_csv(
                content=content
            )

        if (
            file_format
            == ImportFileFormat.XLS
        ):
            return self.parse_xls(
                content=content,
                sheet_name=sheet_name,
            )

        raise ValueError(
            "Непідтримуваний формат."
        )

    # ==========================================
    # XLSX
    # ==========================================

    def parse_xlsx(
        self,
        *,
        content: bytes,
        sheet_name: str | None,
    ) -> ParsedTable:
        """
        Читає XLSX через openpyxl.
        """

        try:
            workbook = load_workbook(
                filename=BytesIO(content),
                read_only=True,
                data_only=True,
            )

        except Exception as error:
            raise ValueError(
                "Не вдалося відкрити XLSX."
            ) from error

        try:
            if sheet_name:
                if (
                    sheet_name
                    not in workbook.sheetnames
                ):
                    raise ValueError(
                        "Аркуш "
                        f"«{sheet_name}» "
                        "не знайдено."
                    )

                worksheet = workbook[
                    sheet_name
                ]

            else:
                worksheet = workbook.active

            raw_rows = [
                tuple(row)
                for row
                in worksheet.iter_rows(
                    values_only=True
                )
            ]

            (
                header_index,
                headers,
            ) = self.find_header_row(
                raw_rows
            )

            rows: list[
                tuple[
                    int,
                    tuple[Any, ...],
                ]
            ] = []

            for index in range(
                header_index + 1,
                len(raw_rows),
            ):
                row = raw_rows[index]

                if self.row_is_empty(
                    row
                ):
                    continue

                rows.append(
                    (
                        index + 1,
                        row,
                    )
                )

                if (
                    len(rows)
                    >= self.MAX_ROWS
                ):
                    break

            return ParsedTable(
                file_format=(
                    ImportFileFormat.XLSX
                ),

                sheet_name=(
                    worksheet.title
                ),

                header_row_number=(
                    header_index + 1
                ),

                headers=tuple(
                    str(value or "")
                    for value in headers
                ),

                rows=tuple(rows),
            )

        finally:
            workbook.close()

    # ==========================================
    # CSV
    # ==========================================

    def parse_csv(
        self,
        *,
        content: bytes,
    ) -> ParsedTable:
        """
        Читає CSV.
        """

        text = self.decode_csv(
            content
        )

        sample = text[
            :10000
        ]

        try:
            dialect = csv.Sniffer().sniff(
                sample,
                delimiters=",;\t|",
            )

        except csv.Error:
            dialect = csv.excel

            if ";" in sample:
                dialect.delimiter = ";"

        reader = csv.reader(
            StringIO(text),
            dialect,
        )

        raw_rows = [
            tuple(row)
            for row in reader
        ]

        (
            header_index,
            headers,
        ) = self.find_header_row(
            raw_rows
        )

        rows: list[
            tuple[
                int,
                tuple[Any, ...],
            ]
        ] = []

        for index in range(
            header_index + 1,
            len(raw_rows),
        ):
            row = raw_rows[index]

            if self.row_is_empty(
                row
            ):
                continue

            rows.append(
                (
                    index + 1,
                    row,
                )
            )

            if (
                len(rows)
                >= self.MAX_ROWS
            ):
                break

        return ParsedTable(
            file_format=(
                ImportFileFormat.CSV
            ),

            sheet_name=None,

            header_row_number=(
                header_index + 1
            ),

            headers=tuple(
                str(value or "")
                for value in headers
            ),

            rows=tuple(rows),
        )

    # ==========================================
    # XLS
    # ==========================================

    def parse_xls(
        self,
        *,
        content: bytes,
        sheet_name: str | None,
    ) -> ParsedTable:
        """
        Читає старий XLS.

        Для цього потрібен xlrd.
        """

        try:
            import xlrd  # type: ignore

        except ImportError as error:
            raise RuntimeError(
                "Для старих .xls файлів "
                "потрібен пакет xlrd. "
                "Або збережіть файл як XLSX."
            ) from error

        try:
            workbook = xlrd.open_workbook(
                file_contents=content
            )

        except Exception as error:
            raise ValueError(
                "Не вдалося відкрити XLS."
            ) from error

        if sheet_name:
            try:
                worksheet = (
                    workbook.sheet_by_name(
                        sheet_name
                    )
                )

            except Exception as error:
                raise ValueError(
                    "Не знайдено аркуш "
                    f"«{sheet_name}»."
                ) from error

        else:
            worksheet = (
                workbook.sheet_by_index(
                    0
                )
            )

        raw_rows = [
            tuple(
                worksheet.row_values(index)
            )
            for index
            in range(
                worksheet.nrows
            )
        ]

        (
            header_index,
            headers,
        ) = self.find_header_row(
            raw_rows
        )

        rows: list[
            tuple[
                int,
                tuple[Any, ...],
            ]
        ] = []

        for index in range(
            header_index + 1,
            len(raw_rows),
        ):
            row = raw_rows[index]

            if self.row_is_empty(
                row
            ):
                continue

            rows.append(
                (
                    index + 1,
                    row,
                )
            )

            if (
                len(rows)
                >= self.MAX_ROWS
            ):
                break

        return ParsedTable(
            file_format=(
                ImportFileFormat.XLS
            ),

            sheet_name=(
                worksheet.name
            ),

            header_row_number=(
                header_index + 1
            ),

            headers=tuple(
                str(value or "")
                for value in headers
            ),

            rows=tuple(rows),
        )

    # ==========================================
    # HEADER DETECTION
    # ==========================================

    def find_header_row(
        self,
        rows: list[
            tuple[Any, ...]
        ],
    ) -> tuple[
        int,
        tuple[Any, ...],
    ]:
        """
        Автоматично знаходить рядок заголовків.
        """

        best_index: int | None = None
        best_score = 0

        scan_limit = min(
            len(rows),
            self.HEADER_SCAN_ROWS,
        )

        for index in range(
            scan_limit
        ):
            row = rows[index]

            mapping = (
                self.resolve_header_mapping(
                    tuple(
                        str(value or "")
                        for value in row
                    )
                )
            )

            score = len(
                mapping
            )

            identifier_found = (
                "code" in mapping
                or "store_number"
                in mapping
            )

            if (
                identifier_found
                and score > best_score
            ):
                best_score = score
                best_index = index

        if best_index is None:
            raise ValueError(
                "Не вдалося знайти "
                "рядок із заголовками."
            )

        return (
            best_index,
            rows[best_index],
        )

    def resolve_header_mapping(
        self,
        headers: tuple[str, ...],
    ) -> dict[str, int]:
        """
        canonical_field -> column index
        """

        mapping: dict[
            str,
            int,
        ] = {}

        for index, header in enumerate(
            headers
        ):
            normalized = (
                self.normalize_header(
                    header
                )
            )

            canonical = (
                self.header_lookup.get(
                    normalized
                )
            )

            if (
                canonical
                and canonical
                not in mapping
            ):
                mapping[
                    canonical
                ] = index

        return mapping

    def map_row_values(
        self,
        *,
        raw_row: tuple[Any, ...],
        header_mapping: dict[
            str,
            int,
        ],
    ) -> dict[str, Any]:
        """
        Витягує canonical values.
        """

        result: dict[
            str,
            Any,
        ] = {}

        for (
            field_name,
            column_index,
        ) in header_mapping.items():
            if (
                column_index
                < len(raw_row)
            ):
                result[field_name] = (
                    raw_row[
                        column_index
                    ]
                )

            else:
                result[
                    field_name
                ] = None

        return result

    # ==========================================
    # REFERENCE MAPS
    # ==========================================

    async def load_existing_stores(
        self,
    ) -> list[Store]:
        """
        Завантажує всі ТТ,
        включно з неактивними.
        """

        result = await self.session.scalars(
            select(Store)
        )

        return list(
            result.unique().all()
        )

    async def load_bushes(
        self,
    ) -> list[Bush]:
        """
        Завантажує кущі.
        """

        result = await self.session.scalars(
            select(Bush)
        )

        return list(
            result.unique().all()
        )

    async def load_clusters(
        self,
    ) -> list[Cluster]:
        """
        Завантажує кластери.
        """

        result = await self.session.scalars(
            select(Cluster)
        )

        return list(
            result.unique().all()
        )

    def build_bush_maps(
        self,
        bushes: list[Bush],
    ) -> dict[str, dict[Any, Bush]]:
        """
        Індекси кущів.
        """

        by_id: dict[Any, Bush] = {}
        by_text: dict[Any, Bush] = {}

        for bush in bushes:
            by_id[bush.id] = bush

            for field_name in (
                "name",
                "title",
                "code",
                "slug",
            ):
                value = getattr(
                    bush,
                    field_name,
                    None,
                )

                if value:
                    by_text[
                        self.normalize_lookup(
                            value
                        )
                    ] = bush

        return {
            "id": by_id,
            "text": by_text,
        }

    def build_cluster_maps(
        self,
        clusters: list[Cluster],
    ) -> dict[str, dict[Any, Cluster]]:
        """
        Індекси кластерів.
        """

        by_id: dict[Any, Cluster] = {}
        by_text: dict[Any, Cluster] = {}
        by_hour: dict[Any, Cluster] = {}

        for cluster in clusters:
            by_id[
                cluster.id
            ] = cluster

            for field_name in (
                "name",
                "title",
                "code",
                "slug",
            ):
                value = getattr(
                    cluster,
                    field_name,
                    None,
                )

                if value:
                    by_text[
                        self.normalize_lookup(
                            value
                        )
                    ] = cluster

            cluster_time = (
                self.cluster_time(
                    cluster
                )
            )

            if cluster_time:
                by_hour[
                    cluster_time.hour
                ] = cluster

                by_text[
                    cluster_time.strftime(
                        "%H:%M"
                    )
                ] = cluster

                by_text[
                    str(
                        cluster_time.hour
                    )
                ] = cluster

        return {
            "id": by_id,
            "text": by_text,
            "hour": by_hour,
        }

    # ==========================================
    # RESOLVE BUSH
    # ==========================================

    def resolve_bush(
        self,
        value: Any,
        maps: dict[
            str,
            dict[Any, Bush],
        ],
    ) -> int | None:
        """
        Розпізнає кущ за ID,
        назвою або кодом.
        """

        normalized = (
            self.normalize_lookup(
                value
            )
        )

        if not normalized:
            return None

        if normalized.isdigit():
            bush = maps[
                "id"
            ].get(
                int(normalized)
            )

            if bush is not None:
                return bush.id

        bush = maps[
            "text"
        ].get(
            normalized
        )

        if bush is not None:
            return bush.id

        return None

    # ==========================================
    # RESOLVE CLUSTER
    # ==========================================

    def resolve_cluster(
        self,
        value: Any,
        maps: dict[
            str,
            dict[Any, Cluster],
        ],
    ) -> int | None:
        """
        Розпізнає кластер:

            8
            08:00
            8:00
            Кластер 08:00
            CLUSTER_08
        """

        normalized = (
            self.normalize_lookup(
                value
            )
        )

        if not normalized:
            return None

        direct = maps[
            "text"
        ].get(
            normalized
        )

        if direct is not None:
            return direct.id

        hour = (
            self.parse_hour_reference(
                normalized
            )
        )

        if hour is not None:
            cluster = maps[
                "hour"
            ].get(
                hour
            )

            if cluster is not None:
                return cluster.id

        if normalized.startswith(
            "id:"
        ):
            id_text = (
                normalized[3:]
                .strip()
            )

            if id_text.isdigit():
                cluster = maps[
                    "id"
                ].get(
                    int(id_text)
                )

                if cluster is not None:
                    return cluster.id

        return None

    # ==========================================
    # UPDATE COMPARISON
    # ==========================================

    def store_needs_update(
        self,
        *,
        store: Store,

        store_number: int | None,
        code: str | None,
        name: str | None,
        city: str | None,
        address: str | None,

        bush_id: int | None,
        cluster_id: int | None,

        is_active: bool | None,
    ) -> bool:
        """
        Чи відрізняється файл від БД.
        """

        if (
            store_number is not None
            and self.store_number(
                store
            )
            != store_number
        ):
            return True

        if (
            code is not None
            and self.store_code(
                store
            ).lower()
            != code.lower()
        ):
            return True

        if (
            name is not None
            and self.store_name(
                store
            )
            != name
        ):
            return True

        if (
            city is not None
            and (
                self.get_text_attribute(
                    store,
                    "city",
                )
                or ""
            )
            != city
        ):
            return True

        if (
            (
                self.get_text_attribute(
                    store,
                    "address",
                )
                or None
            )
            != address
        ):
            return True

        if (
            getattr(
                store,
                "bush_id",
                None,
            )
            != bush_id
        ):
            return True

        if (
            getattr(
                store,
                "cluster_id",
                None,
            )
            != cluster_id
        ):
            return True

        if (
            is_active is not None
            and bool(
                getattr(
                    store,
                    "is_active",
                    True,
                )
            )
            != is_active
        ):
            return True

        return False

    # ==========================================
    # AUDIT
    # ==========================================

    async def log_import(
        self,
        *,
        actor: User,
        preview: ImportPreview,
        result: ImportApplyResult,
        reason: str,
    ) -> None:
        """
        Записує підсумок імпорту.
        """

        action = (
            self.resolve_audit_action(
                "update",
                "import",
                "changed",
            )
        )

        entity_type = (
            self.resolve_entity_type(
                "system",
                "store",
            )
        )

        await self.repositories.audit.log_action(
            action=action,

            entity_type=entity_type,

            entity_id=None,

            context=AuditContext(
                actor_user_id=actor.id,
                reason=reason,
                description=(
                    "Імпорт торгових точок"
                ),
                source="telegram_bot",
            ),

            old_values={
                "file_name": (
                    preview.file_name
                ),
                "preview_rows": (
                    preview.total_rows
                ),
            },

            new_values={
                "attempted": (
                    result.attempted_count
                ),
                "success": (
                    result.success_count
                ),
                "failed": (
                    result.failed_count
                ),
                "created": (
                    result.created_count
                ),
                "updated": (
                    result.updated_count
                ),
                "ignored": (
                    result.ignored_count
                ),
                "invalid": (
                    result.invalid_count
                ),
            },
        )

    # ==========================================
    # FILE FORMAT
    # ==========================================

    @staticmethod
    def detect_file_format(
        file_name: str,
    ) -> ImportFileFormat:
        """
        Визначає формат за extension.
        """

        extension = Path(
            file_name
        ).suffix.lower()

        mapping = {
            ".xlsx": (
                ImportFileFormat.XLSX
            ),
            ".xls": (
                ImportFileFormat.XLS
            ),
            ".csv": (
                ImportFileFormat.CSV
            ),
        }

        result = mapping.get(
            extension
        )

        if result is None:
            raise ValueError(
                "Дозволені файли: "
                "XLSX, XLS або CSV."
            )

        return result

    # ==========================================
    # CSV ENCODING
    # ==========================================

    @staticmethod
    def decode_csv(
        content: bytes,
    ) -> str:
        """
        Підбирає кодування CSV.
        """

        encodings = (
            "utf-8-sig",
            "utf-8",
            "cp1251",
            "windows-1251",
        )

        for encoding in encodings:
            try:
                return content.decode(
                    encoding
                )

            except UnicodeDecodeError:
                continue

        raise ValueError(
            "Не вдалося визначити "
            "кодування CSV."
        )

    # ==========================================
    # HEADER NORMALIZATION
    # ==========================================

    def build_header_lookup(
        self,
    ) -> dict[str, str]:
        """
        alias -> canonical field.
        """

        result: dict[
            str,
            str,
        ] = {}

        for (
            canonical,
            aliases,
        ) in self.HEADER_ALIASES.items():
            result[
                self.normalize_header(
                    canonical
                )
            ] = canonical

            for alias in aliases:
                result[
                    self.normalize_header(
                        alias
                    )
                ] = canonical

        return result

    @staticmethod
    def normalize_header(
        value: Any,
    ) -> str:
        """
        Нормалізує заголовок.
        """

        normalized = str(
            value or ""
        ).strip().lower()

        normalized = normalized.replace(
            "\n",
            " ",
        )

        normalized = normalized.replace(
            "_",
            " ",
        )

        normalized = re.sub(
            r"\s+",
            " ",
            normalized,
        )

        return normalized.strip()

    # ==========================================
    # STORE NUMBER
    # ==========================================

    @staticmethod
    def parse_store_number(
        value: Any,
    ) -> int | None:
        """
        Розпізнає номер ТТ.
        """

        if value is None:
            return None

        if isinstance(
            value,
            bool,
        ):
            return None

        if isinstance(
            value,
            int,
        ):
            return (
                value
                if value > 0
                else None
            )

        if isinstance(
            value,
            float,
        ):
            if value.is_integer():
                integer = int(value)

                return (
                    integer
                    if integer > 0
                    else None
                )

        text = str(
            value
        ).strip()

        if not text:
            return None

        numbers = re.findall(
            r"\d+",
            text,
        )

        if not numbers:
            return None

        number = int(
            numbers[-1]
        )

        return (
            number
            if number > 0
            else None
        )

    @staticmethod
    def number_from_code(
        code: str,
    ) -> int | None:
        """
        SB-76 -> 76
        """

        match = re.search(
            r"(\d+)$",
            code,
        )

        if not match:
            return None

        return int(
            match.group(1)
        )

    # ==========================================
    # BOOL
    # ==========================================

    @staticmethod
    def parse_bool(
        value: Any,
    ) -> bool:
        """
        Розпізнає статус.
        """

        if isinstance(
            value,
            bool,
        ):
            return value

        if isinstance(
            value,
            (int, float),
        ):
            if value == 1:
                return True

            if value == 0:
                return False

        normalized = str(
            value
        ).strip().lower()

        true_values = {
            "1",
            "true",
            "yes",
            "y",
            "так",
            "да",
            "active",
            "активний",
            "активна",
            "працює",
            "работает",
            "відкритий",
        }

        false_values = {
            "0",
            "false",
            "no",
            "n",
            "ні",
            "нет",
            "inactive",
            "неактивний",
            "неактивна",
            "закритий",
            "закрыт",
            "не працює",
        }

        if normalized in true_values:
            return True

        if normalized in false_values:
            return False

        raise ValueError(
            f"Невідомий статус: {value}."
        )

    # ==========================================
    # KYIV
    # ==========================================

    @classmethod
    def is_kyiv(
        cls,
        value: str,
    ) -> bool:
        """
        Визначає Київ.
        """

        normalized = (
            cls.normalize_lookup(
                value
            )
        )

        normalized = re.sub(
            r"^(м|місто|город)[.\s]+",
            "",
            normalized,
        )

        if normalized in cls.KYIV_NAMES:
            return True

        words = set(
            re.findall(
                r"[a-zа-яіїєґ]+",
                normalized,
            )
        )

        return bool(
            words.intersection(
                cls.KYIV_NAMES
            )
        )

    # ==========================================
    # CLUSTER TIME
    # ==========================================

    @staticmethod
    def cluster_time(
        cluster: Cluster,
    ) -> time | None:
        """
        Витягує час кластера.
        """

        for field_name in (
            "opening_time",
            "start_time",
        ):
            value = getattr(
                cluster,
                field_name,
                None,
            )

            if isinstance(
                value,
                time,
            ):
                return time(
                    value.hour,
                    value.minute,
                )

        for field_name in (
            "hour",
            "opening_hour",
        ):
            value = getattr(
                cluster,
                field_name,
                None,
            )

            if value is None:
                continue

            try:
                hour = int(
                    value
                )

            except (
                TypeError,
                ValueError,
            ):
                continue

            if 0 <= hour <= 23:
                return time(
                    hour,
                    0,
                )

        return None

    @staticmethod
    def parse_hour_reference(
        value: str,
    ) -> int | None:
        """
        Витягує годину з:

            8
            08:00
            кластер 08:00
            8.00
        """

        matches = re.findall(
            r"(?<!\d)(\d{1,2})(?::|\.|$)",
            value,
        )

        for match in matches:
            try:
                hour = int(
                    match
                )

            except ValueError:
                continue

            if 0 <= hour <= 23:
                return hour

        return None

    # ==========================================
    # STORE HELPERS
    # ==========================================

    @staticmethod
    def store_number(
        store: Store,
    ) -> int | None:
        """
        Номер ТТ.
        """

        for field_name in (
            "store_number",
            "number",
        ):
            value = getattr(
                store,
                field_name,
                None,
            )

            if value is None:
                continue

            try:
                return int(
                    value
                )

            except (
                TypeError,
                ValueError,
            ):
                continue

        code = getattr(
            store,
            "code",
            None,
        )

        if code:
            return (
                ImportService
                .number_from_code(
                    str(code)
                )
            )

        return None

    @classmethod
    def store_code(
        cls,
        store: Store,
    ) -> str:
        """
        Код ТТ.
        """

        code = getattr(
            store,
            "code",
            None,
        )

        if code:
            return str(code)

        number = cls.store_number(
            store
        )

        if number is not None:
            return f"SB-{number}"

        return f"ТТ-{store.id}"

    @classmethod
    def store_name(
        cls,
        store: Store | None,
    ) -> str | None:
        """
        Назва ТТ.
        """

        if store is None:
            return None

        for field_name in (
            "name",
            "title",
            "display_name",
        ):
            value = getattr(
                store,
                field_name,
                None,
            )

            if value:
                return str(value)

        return cls.store_code(
            store
        )

    # ==========================================
    # GENERIC HELPERS
    # ==========================================

    @staticmethod
    def get_text_attribute(
        target: Any,
        *names: str,
    ) -> str | None:
        """
        Читає текстове поле.
        """

        if target is None:
            return None

        for name in names:
            value = getattr(
                target,
                name,
                None,
            )

            if value is None:
                continue

            normalized = str(
                value
            ).strip()

            if normalized:
                return normalized

        return None

    @staticmethod
    def has_value(
        value: Any,
    ) -> bool:
        """
        Чи не порожня клітинка.
        """

        if value is None:
            return False

        if isinstance(
            value,
            str,
        ):
            return bool(
                value.strip()
            )

        return True

    @classmethod
    def row_is_empty(
        cls,
        row: tuple[Any, ...],
    ) -> bool:
        """
        Чи повністю порожній рядок.
        """

        return not any(
            cls.has_value(
                value
            )
            for value in row
        )

    @staticmethod
    def normalize_lookup(
        value: Any,
    ) -> str:
        """
        Нормалізує значення для пошуку.
        """

        normalized = str(
            value or ""
        ).strip().lower()

        normalized = (
            normalized
            .replace("’", "'")
            .replace("`", "'")
        )

        normalized = re.sub(
            r"\s+",
            " ",
            normalized,
        )

        return normalized.strip()

    @staticmethod
    def normalize_optional_text(
        value: Any,
    ) -> str | None:
        """
        Необов’язковий текст.
        """

        if value is None:
            return None

        normalized = " ".join(
            str(value)
            .strip()
            .split()
        )

        return normalized or None

    @staticmethod
    def normalize_required_text(
        value: str,
        *,
        field_name: str,
        max_length: int,
    ) -> str:
        """
        Обов’язковий текст.
        """

        normalized = " ".join(
            value.strip().split()
        )

        if not normalized:
            raise ValueError(
                f"{field_name} "
                "не може бути порожнім."
            )

        if len(normalized) > max_length:
            raise ValueError(
                f"{field_name} "
                "занадто довгий."
            )

        return normalized

    # ==========================================
    # ENUM
    # ==========================================

    @classmethod
    def resolve_audit_action(
        cls,
        *names: str,
    ) -> AuditAction:
        """
        AuditAction.
        """

        result = cls.resolve_enum_member(
            AuditAction,
            *names,
            default=None,
        )

        if result is not None:
            return result

        result = cls.resolve_enum_member(
            AuditAction,
            "update",
            "changed",
            default=None,
        )

        if result is None:
            raise ValueError(
                "Не знайдено AuditAction."
            )

        return result

    @classmethod
    def resolve_entity_type(
        cls,
        *names: str,
    ) -> EntityType:
        """
        EntityType.
        """

        result = cls.resolve_enum_member(
            EntityType,
            *names,
            default=None,
        )

        if result is not None:
            return result

        result = cls.resolve_enum_member(
            EntityType,
            "system",
            default=None,
        )

        if result is None:
            raise ValueError(
                "Не знайдено EntityType."
            )

        return result

    @staticmethod
    def resolve_enum_member(
        enum_class: type[EnumType],
        *names: str,
        default: EnumType | None = None,
    ) -> EnumType | None:
        """
        Пошук enum.
        """

        normalized = {
            name.strip().lower()
            for name in names
            if name.strip()
        }

        for item in enum_class:
            candidates = {
                item.name.lower(),
                str(
                    item.value
                ).lower(),
            }

            if candidates.intersection(
                normalized
            ):
                return item

        return default

    # ==========================================
    # TELEGRAM FORMAT
    # ==========================================

    @staticmethod
    def format_preview(
        preview: ImportPreview,
    ) -> str:
        """
        Preview для Telegram.
        """

        lines = [
            "📥 <b>Preview імпорту ТТ</b>",
            "",
            (
                "Файл: "
                f"<code>{preview.file_name}</code>"
            ),
            (
                "Рядків: "
                f"<b>{preview.total_rows}</b>"
            ),
            "",
            (
                "🆕 Нових: "
                f"<b>{preview.create_count}</b>"
            ),
            (
                "🔄 Оновлення: "
                f"<b>{preview.update_count}</b>"
            ),
            (
                "✅ Без змін: "
                f"<b>{preview.unchanged_count}</b>"
            ),
            (
                "🚫 Пропущено: "
                f"<b>{preview.ignored_count}</b>"
            ),
            (
                "❌ Помилок: "
                f"<b>{preview.invalid_count}</b>"
            ),
        ]

        if preview.invalid_count:
            lines.extend(
                [
                    "",
                    "⚠️ <b>Проблемні рядки:</b>",
                ]
            )

            shown = 0

            for row in preview.rows:
                if (
                    row.status
                    != ImportRowStatus.INVALID
                ):
                    continue

                errors = [
                    issue.message
                    for issue in row.issues
                    if (
                        issue.level
                        == ImportIssueLevel.ERROR
                    )
                ]

                lines.append(
                    f"• Рядок {row.row_number}: "
                    + "; ".join(
                        errors
                    )
                )

                shown += 1

                if shown >= 10:
                    break

        elif preview.actionable_count:
            lines.extend(
                [
                    "",
                    "✅ <b>Файл готовий до імпорту.</b>",
                ]
            )

        else:
            lines.extend(
                [
                    "",
                    "ℹ️ Змін для запису немає.",
                ]
            )

        return "\n".join(
            lines
        )

    @staticmethod
    def format_apply_result(
        result: ImportApplyResult,
    ) -> str:
        """
        Результат імпорту для Telegram.
        """

        lines = [
            "✅ <b>Імпорт завершено</b>",
            "",
            (
                "Оброблено: "
                f"<b>{result.attempted_count}</b>"
            ),
            (
                "Успішно: "
                f"<b>{result.success_count}</b>"
            ),
            (
                "Помилки: "
                f"<b>{result.failed_count}</b>"
            ),
            "",
            (
                "🆕 Створено: "
                f"<b>{result.created_count}</b>"
            ),
            (
                "🔄 Оновлено: "
                f"<b>{result.updated_count}</b>"
            ),
            (
                "✅ Без змін: "
                f"<b>{result.unchanged_count}</b>"
            ),
            (
                "🚫 Пропущено: "
                f"<b>{result.ignored_count}</b>"
            ),
        ]

        if result.failed_count:
            lines.extend(
                [
                    "",
                    "⚠️ <b>Помилки:</b>",
                ]
            )

            failed = [
                item
                for item in result.items
                if not item.success
            ]

            for item in failed[:10]:
                lines.append(
                    "• Рядок "
                    f"{item.row_number}: "
                    f"{item.error or 'невідома помилка'}"
                )

        return "\n".join(
            lines
        )


__all__ = [
    "ImportService",
    "ImportFileFormat",
    "ImportRowStatus",
    "ImportIssueLevel",
    "ImportIssue",
    "StoreImportRow",
    "ImportPreview",
    "ImportApplyItemResult",
    "ImportApplyResult",
    "ParsedTable",
]