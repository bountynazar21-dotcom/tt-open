from __future__ import annotations

import json
import re
from dataclasses import dataclass
from datetime import date, datetime, time
from decimal import Decimal, InvalidOperation
from enum import Enum
from io import BytesIO
from pathlib import Path
from typing import Any

from aiogram.types import BufferedInputFile
from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from app.services.report_service import (
    ExcelReportData,
    ExcelSheetData,
)


@dataclass(slots=True, frozen=True)
class GeneratedExcelFile:
    """
    Готовий Excel-файл у пам’яті.
    """

    filename: str
    content: bytes

    size_bytes: int
    workbook_title: str

    penalty_per_minute: Decimal

    @property
    def size_kb(self) -> float:
        return round(
            self.size_bytes / 1024,
            2,
        )

    def as_telegram_file(
        self,
    ) -> BufferedInputFile:
        """
        Перетворює файл у формат для Aiogram.
        """

        return BufferedInputFile(
            self.content,
            filename=self.filename,
        )


@dataclass(slots=True, frozen=True)
class PreparedSheetData:
    """
    Підготовлені дані Excel-аркуша.

    Тут уже додані штрафні колонки.
    """

    title: str

    headers: tuple[str, ...]

    rows: tuple[
        tuple[Any, ...],
        ...,
    ]

    column_widths: tuple[float, ...]

    freeze_panes: str
    auto_filter: bool


class ExcelService:
    """
    Сервіс фізичного створення Excel-файлів.

    Підтримує:

    - фіолетовий стиль Soska Bar;
    - автоматичні фільтри;
    - закріплений заголовок;
    - форматування дат;
    - форматування часу;
    - форматування каси;
    - автоматичну ширину колонок;
    - підсвічування проблемних ТТ;
    - формування файлу для Telegram;
    - збереження файлу на диск;
    - автоматичний розрахунок штрафу.

    Правило штрафу:

        1 хвилина запізнення = 8 грн.
    """

    PENALTY_PER_MINUTE = Decimal("8.00")

    # ==========================================
    # ФІРМОВІ КОЛЬОРИ
    # ==========================================

    COLOR_PURPLE = "7C3AED"
    COLOR_PURPLE_DARK = "4C1D95"
    COLOR_PURPLE_LIGHT = "EDE9FE"

    COLOR_WHITE = "FFFFFF"
    COLOR_BLACK = "111827"

    COLOR_GRAY_LIGHT = "F3F4F6"
    COLOR_GRAY = "6B7280"

    COLOR_GREEN_LIGHT = "DCFCE7"
    COLOR_GREEN_TEXT = "166534"

    COLOR_RED_LIGHT = "FEE2E2"
    COLOR_RED_TEXT = "991B1B"

    COLOR_YELLOW_LIGHT = "FEF3C7"
    COLOR_YELLOW_TEXT = "92400E"

    COLOR_BLUE_LIGHT = "DBEAFE"
    COLOR_BLUE_TEXT = "1E40AF"

    # ==========================================
    # СТИЛІ
    # ==========================================

    THIN_BORDER = Border(
        left=Side(
            style="thin",
            color="D1D5DB",
        ),
        right=Side(
            style="thin",
            color="D1D5DB",
        ),
        top=Side(
            style="thin",
            color="D1D5DB",
        ),
        bottom=Side(
            style="thin",
            color="D1D5DB",
        ),
    )

    HEADER_FILL = PatternFill(
        fill_type="solid",
        fgColor=COLOR_PURPLE,
    )

    TITLE_FILL = PatternFill(
        fill_type="solid",
        fgColor=COLOR_PURPLE_DARK,
    )

    LIGHT_PURPLE_FILL = PatternFill(
        fill_type="solid",
        fgColor=COLOR_PURPLE_LIGHT,
    )

    WHITE_FONT = Font(
        name="Arial",
        size=11,
        bold=True,
        color=COLOR_WHITE,
    )

    HEADER_FONT = Font(
        name="Arial",
        size=10,
        bold=True,
        color=COLOR_WHITE,
    )

    BODY_FONT = Font(
        name="Arial",
        size=10,
        color=COLOR_BLACK,
    )

    # ==========================================
    # ГЕНЕРАЦІЯ ФАЙЛУ
    # ==========================================

    def generate(
        self,
        report_data: ExcelReportData,
    ) -> GeneratedExcelFile:
        """
        Створює готовий .xlsx-файл у пам’яті.
        """

        self.validate_report_data(
            report_data
        )

        workbook = self.build_workbook(
            report_data
        )

        buffer = BytesIO()

        workbook.save(buffer)

        content = buffer.getvalue()

        buffer.close()

        filename = self.sanitize_filename(
            report_data.filename
        )

        return GeneratedExcelFile(
            filename=filename,
            content=content,
            size_bytes=len(content),
            workbook_title=(
                report_data.workbook_title
            ),
            penalty_per_minute=(
                self.PENALTY_PER_MINUTE
            ),
        )

    def generate_telegram_file(
        self,
        report_data: ExcelReportData,
    ) -> BufferedInputFile:
        """
        Одразу повертає файл для bot.send_document().
        """

        generated = self.generate(
            report_data
        )

        return generated.as_telegram_file()

    def save(
        self,
        report_data: ExcelReportData,
        *,
        directory: str | Path,
    ) -> Path:
        """
        Створює Excel і зберігає його на диск.
        """

        generated = self.generate(
            report_data
        )

        target_directory = Path(
            directory
        )

        target_directory.mkdir(
            parents=True,
            exist_ok=True,
        )

        file_path = (
            target_directory
            / generated.filename
        )

        file_path.write_bytes(
            generated.content
        )

        return file_path

    # ==========================================
    # СТВОРЕННЯ WORKBOOK
    # ==========================================

    def build_workbook(
        self,
        report_data: ExcelReportData,
    ) -> Workbook:
        """
        Створює openpyxl Workbook.
        """

        workbook = Workbook()

        default_sheet = workbook.active

        workbook.remove(default_sheet)

        workbook.properties.title = (
            report_data.workbook_title
        )

        workbook.properties.subject = (
            "Звіт Telegram-бота контролю ТТ"
        )

        workbook.properties.creator = (
            "Soska Bar"
        )

        workbook.properties.company = (
            "Soska Bar"
        )

        used_titles: set[str] = set()

        for sheet_data in report_data.sheets:
            prepared_data = (
                self.prepare_sheet_data(
                    sheet_data
                )
            )

            sheet_title = self.unique_sheet_title(
                prepared_data.title,
                used_titles=used_titles,
            )

            used_titles.add(sheet_title)

            worksheet = workbook.create_sheet(
                title=sheet_title
            )

            self.write_sheet(
                worksheet=worksheet,
                sheet_data=prepared_data,
                workbook_title=(
                    report_data.workbook_title
                ),
            )

        self.create_metadata_sheet(
            workbook=workbook,
            report_data=report_data,
        )

        if workbook.worksheets:
            workbook.active = 0

        return workbook

    # ==========================================
    # ПІДГОТОВКА ДАНИХ АРКУША
    # ==========================================

    def prepare_sheet_data(
        self,
        sheet_data: ExcelSheetData,
    ) -> PreparedSheetData:
        """
        Додає штрафи до звітних даних.

        Підтримує два типи таблиць:

        1. Звичайна таблиця:
           є колонка «Запізнення, хв».

        2. Таблиця підсумків:
           є рядок «Загальне запізнення, хв».
        """

        headers = list(
            sheet_data.headers
        )

        rows = [
            list(row)
            for row in sheet_data.rows
        ]

        widths = list(
            sheet_data.column_widths
        )

        if self.is_metrics_sheet(headers):
            rows = self.add_penalty_metric_rows(
                rows
            )

        else:
            (
                headers,
                rows,
                widths,
            ) = self.add_penalty_columns(
                headers=headers,
                rows=rows,
                widths=widths,
            )

        self.ensure_widths_count(
            widths=widths,
            columns_count=len(headers),
        )

        return PreparedSheetData(
            title=sheet_data.title,
            headers=tuple(headers),
            rows=tuple(
                tuple(row)
                for row in rows
            ),
            column_widths=tuple(widths),
            freeze_panes=(
                sheet_data.freeze_panes
            ),
            auto_filter=(
                sheet_data.auto_filter
            ),
        )

    @staticmethod
    def is_metrics_sheet(
        headers: list[str],
    ) -> bool:
        """
        Чи є аркуш таблицею показник-значення.
        """

        if len(headers) < 2:
            return False

        first_header = (
            str(headers[0])
            .strip()
            .lower()
        )

        second_header = (
            str(headers[1])
            .strip()
            .lower()
        )

        return (
            first_header == "показник"
            and second_header == "значення"
        )

    # ==========================================
    # ШТРАФНА КОЛОНКА
    # ==========================================

    def add_penalty_columns(
        self,
        *,
        headers: list[str],
        rows: list[list[Any]],
        widths: list[float],
    ) -> tuple[
        list[str],
        list[list[Any]],
        list[float],
    ]:
        """
        Додає колонку штрафу після запізнення.

        Наприклад:

        Запізнення, хв | Штраф, грн
        5              | 40
        10             | 80
        """

        normalized_headers = [
            str(header).strip().lower()
            for header in headers
        ]

        if any(
            "штраф" in header
            for header in normalized_headers
        ):
            return headers, rows, widths

        penalty_sources = (
            "запізнення, хв",
            "запізнення хв",
            "загальне запізнення, хв",
            "загальне запізнення хв",
        )

        source_index: int | None = None

        for index, header in enumerate(
            normalized_headers
        ):
            if header in penalty_sources:
                source_index = index
                break

        if source_index is None:
            return headers, rows, widths

        source_header = (
            normalized_headers[
                source_index
            ]
        )

        if "загальне" in source_header:
            penalty_header = (
                "Штраф за запізнення, грн"
            )

        else:
            penalty_header = "Штраф, грн"

        insert_index = source_index + 1

        headers.insert(
            insert_index,
            penalty_header,
        )

        if source_index < len(widths):
            widths.insert(
                insert_index,
                20,
            )

        else:
            widths.append(20)

        for row in rows:
            while len(row) < len(headers) - 1:
                row.append(None)

            lateness_minutes = row[
                source_index
            ]

            penalty = self.calculate_penalty(
                lateness_minutes
            )

            row.insert(
                insert_index,
                penalty,
            )

        return headers, rows, widths

    def add_penalty_metric_rows(
        self,
        rows: list[list[Any]],
    ) -> list[list[Any]]:
        """
        Додає штраф окремим рядком у підсумках.

        Загальне запізнення, хв | 20
        Штраф за запізнення, грн | 160
        """

        result: list[list[Any]] = []

        existing_penalty_row = any(
            row
            and "штраф" in str(
                row[0]
            ).lower()
            for row in rows
        )

        for row in rows:
            result.append(row)

            if existing_penalty_row:
                continue

            if not row:
                continue

            metric_name = (
                str(row[0])
                .strip()
                .lower()
            )

            if metric_name not in {
                "загальне запізнення, хв",
                "загальне запізнення хв",
            }:
                continue

            minutes = (
                row[1]
                if len(row) > 1
                else 0
            )

            penalty = self.calculate_penalty(
                minutes
            )

            result.append(
                [
                    (
                        "Штраф за запізнення, грн"
                    ),
                    penalty,
                ]
            )

        return result

    # ==========================================
    # РОЗРАХУНОК ШТРАФУ
    # ==========================================

    @classmethod
    def calculate_penalty(
        cls,
        lateness_minutes: Any,
    ) -> Decimal:
        """
        Рахує штраф.

        Формула:

            хвилини запізнення × 8 грн
        """

        minutes = cls.normalize_minutes(
            lateness_minutes
        )

        return (
            Decimal(minutes)
            * cls.PENALTY_PER_MINUTE
        ).quantize(
            Decimal("0.01")
        )

    @staticmethod
    def normalize_minutes(
        value: Any,
    ) -> int:
        """
        Нормалізує хвилини запізнення.
        """

        if value is None:
            return 0

        if isinstance(value, bool):
            return 0

        if isinstance(value, Decimal):
            number = int(value)

        elif isinstance(
            value,
            (int, float),
        ):
            number = int(value)

        else:
            normalized_value = re.sub(
                r"[^0-9\-]",
                "",
                str(value),
            )

            if not normalized_value:
                return 0

            try:
                number = int(
                    normalized_value
                )

            except ValueError:
                return 0

        return max(number, 0)

    # ==========================================
    # ЗАПИС АРКУША
    # ==========================================

    def write_sheet(
        self,
        *,
        worksheet: Worksheet,
        sheet_data: PreparedSheetData,
        workbook_title: str,
    ) -> None:
        """
        Записує повністю оформлений аркуш.
        """

        columns_count = max(
            len(sheet_data.headers),
            1,
        )

        last_column_letter = (
            get_column_letter(
                columns_count
            )
        )

        # --------------------------------------
        # ГОЛОВНИЙ ЗАГОЛОВОК
        # --------------------------------------

        worksheet.merge_cells(
            start_row=1,
            start_column=1,
            end_row=1,
            end_column=columns_count,
        )

        title_cell = worksheet.cell(
            row=1,
            column=1,
            value=sheet_data.title,
        )

        title_cell.fill = self.TITLE_FILL

        title_cell.font = Font(
            name="Arial",
            size=16,
            bold=True,
            color=self.COLOR_WHITE,
        )

        title_cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
        )

        worksheet.row_dimensions[1].height = 30

        # --------------------------------------
        # ПІДЗАГОЛОВОК
        # --------------------------------------

        worksheet.merge_cells(
            start_row=2,
            start_column=1,
            end_row=2,
            end_column=columns_count,
        )

        subtitle_cell = worksheet.cell(
            row=2,
            column=1,
            value=(
                f"{workbook_title} · "
                "Штраф за запізнення: "
                f"{self.format_decimal(self.PENALTY_PER_MINUTE)} "
                "грн за 1 хв"
            ),
        )

        subtitle_cell.fill = (
            self.LIGHT_PURPLE_FILL
        )

        subtitle_cell.font = Font(
            name="Arial",
            size=10,
            italic=True,
            color=self.COLOR_PURPLE_DARK,
        )

        subtitle_cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
        )

        worksheet.row_dimensions[2].height = 22

        # --------------------------------------
        # ЗАГОЛОВКИ КОЛОНОК
        # --------------------------------------

        header_row = 3

        for column_index, header in enumerate(
            sheet_data.headers,
            start=1,
        ):
            cell = worksheet.cell(
                row=header_row,
                column=column_index,
                value=header,
            )

            cell.fill = self.HEADER_FILL
            cell.font = self.HEADER_FONT

            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True,
            )

            cell.border = self.THIN_BORDER

        worksheet.row_dimensions[
            header_row
        ].height = 38

        # --------------------------------------
        # ДАНІ
        # --------------------------------------

        first_data_row = header_row + 1

        for row_index, source_row in enumerate(
            sheet_data.rows,
            start=first_data_row,
        ):
            for column_index, value in enumerate(
                source_row,
                start=1,
            ):
                header = (
                    sheet_data.headers[
                        column_index - 1
                    ]
                    if (
                        column_index - 1
                        < len(
                            sheet_data.headers
                        )
                    )
                    else ""
                )

                cell = worksheet.cell(
                    row=row_index,
                    column=column_index,
                    value=self.normalize_excel_value(
                        value
                    ),
                )

                self.style_data_cell(
                    cell=cell,
                    header=header,
                    value=value,
                    row_index=row_index,
                )

        last_data_row = (
            first_data_row
            + len(sheet_data.rows)
            - 1
        )

        # --------------------------------------
        # ФІЛЬТР
        # --------------------------------------

        if (
            sheet_data.auto_filter
            and sheet_data.rows
        ):
            worksheet.auto_filter.ref = (
                f"A{header_row}:"
                f"{last_column_letter}"
                f"{last_data_row}"
            )

        # --------------------------------------
        # ЗАКРІПЛЕННЯ
        # --------------------------------------

        worksheet.freeze_panes = "A4"

        # --------------------------------------
        # ШИРИНА КОЛОНОК
        # --------------------------------------

        self.apply_column_widths(
            worksheet=worksheet,
            widths=(
                sheet_data.column_widths
            ),
            columns_count=columns_count,
        )

        # --------------------------------------
        # УМОВНЕ ФОРМАТУВАННЯ
        # --------------------------------------

        if sheet_data.rows:
            self.apply_conditional_formatting(
                worksheet=worksheet,
                headers=sheet_data.headers,
                first_data_row=first_data_row,
                last_data_row=last_data_row,
            )

        # --------------------------------------
        # НАЛАШТУВАННЯ ДРУКУ
        # --------------------------------------

        worksheet.sheet_view.showGridLines = (
            False
        )

        worksheet.auto_filter.ref = (
            worksheet.auto_filter.ref
        )

        worksheet.print_title_rows = "1:3"

        worksheet.sheet_properties.pageSetUpPr.fitToPage = (
            True
        )

        worksheet.page_setup.fitToWidth = 1
        worksheet.page_setup.fitToHeight = 0
        worksheet.page_setup.orientation = (
            "landscape"
        )

        worksheet.page_margins.left = 0.25
        worksheet.page_margins.right = 0.25
        worksheet.page_margins.top = 0.5
        worksheet.page_margins.bottom = 0.5

        worksheet.oddFooter.center.text = (
            "Soska Bar · Контроль відкриття ТТ"
        )

        worksheet.oddFooter.right.text = (
            "Сторінка &P з &N"
        )

    # ==========================================
    # СТИЛЬ КОМІРКИ
    # ==========================================

    def style_data_cell(
        self,
        *,
        cell: Any,
        header: str,
        value: Any,
        row_index: int,
    ) -> None:
        """
        Застосовує стиль до комірки.
        """

        normalized_header = (
            str(header)
            .strip()
            .lower()
        )

        cell.font = self.BODY_FONT

        cell.border = self.THIN_BORDER

        cell.alignment = Alignment(
            vertical="center",
            horizontal=self.resolve_alignment(
                normalized_header
            ),
            wrap_text=True,
        )

        if row_index % 2 == 0:
            cell.fill = PatternFill(
                fill_type="solid",
                fgColor="FAFAFA",
            )

        self.apply_number_format(
            cell=cell,
            header=normalized_header,
            value=value,
        )

        if "статус" in normalized_header:
            self.style_status_cell(
                cell
            )

        if "фото чека" in normalized_header:
            self.style_yes_no_cell(
                cell
            )

        if "штраф" in normalized_header:
            self.style_penalty_cell(
                cell
            )

    # ==========================================
    # ФОРМАТ ЧИСЕЛ
    # ==========================================

    def apply_number_format(
        self,
        *,
        cell: Any,
        header: str,
        value: Any,
    ) -> None:
        """
        Встановлює Excel number format.
        """

        if isinstance(
            value,
            datetime,
        ):
            cell.number_format = (
                "dd.mm.yyyy hh:mm"
            )
            return

        if isinstance(value, date):
            cell.number_format = (
                "dd.mm.yyyy"
            )
            return

        if isinstance(value, time):
            cell.number_format = "hh:mm"
            return

        if (
            "каса" in header
            or "штраф" in header
        ):
            cell.number_format = (
                '#,##0.00 "грн"'
            )
            return

        if "%" in header:
            cell.number_format = "0.00"
            return

        if "хв" in header:
            cell.number_format = "0"
            return

        if header.endswith("id"):
            cell.number_format = "0"

    # ==========================================
    # КОЛЬОРИ СТАТУСІВ
    # ==========================================

    def style_status_cell(
        self,
        cell: Any,
    ) -> None:
        """
        Підсвічує статус відкриття або закриття.
        """

        value = (
            str(cell.value or "")
            .strip()
            .lower()
        )

        if not value:
            return

        if any(
            marker in value
            for marker in (
                "пропущ",
                "не подано",
                "не відкри",
                "помилка",
            )
        ):
            cell.fill = PatternFill(
                fill_type="solid",
                fgColor=self.COLOR_RED_LIGHT,
            )

            cell.font = Font(
                name="Arial",
                size=10,
                bold=True,
                color=self.COLOR_RED_TEXT,
            )

            return

        if "запіз" in value:
            cell.fill = PatternFill(
                fill_type="solid",
                fgColor=self.COLOR_YELLOW_LIGHT,
            )

            cell.font = Font(
                name="Arial",
                size=10,
                bold=True,
                color=self.COLOR_YELLOW_TEXT,
            )

            return

        if any(
            marker in value
            for marker in (
                "вчасно",
                "відкрито",
                "подано",
                "підтверджено",
            )
        ):
            cell.fill = PatternFill(
                fill_type="solid",
                fgColor=self.COLOR_GREEN_LIGHT,
            )

            cell.font = Font(
                name="Arial",
                size=10,
                bold=True,
                color=self.COLOR_GREEN_TEXT,
            )

            return

        if "очіку" in value:
            cell.fill = PatternFill(
                fill_type="solid",
                fgColor=self.COLOR_BLUE_LIGHT,
            )

            cell.font = Font(
                name="Arial",
                size=10,
                bold=True,
                color=self.COLOR_BLUE_TEXT,
            )

            return

        if any(
            marker in value
            for marker in (
                "вихідний",
                "не потрібен",
                "тимчасово",
                "немає запису",
            )
        ):
            cell.fill = PatternFill(
                fill_type="solid",
                fgColor=self.COLOR_GRAY_LIGHT,
            )

            cell.font = Font(
                name="Arial",
                size=10,
                color=self.COLOR_GRAY,
            )

    def style_yes_no_cell(
        self,
        cell: Any,
    ) -> None:
        """
        Підсвічує Так/Ні.
        """

        normalized_value = (
            str(cell.value or "")
            .strip()
            .lower()
        )

        if normalized_value == "так":
            cell.fill = PatternFill(
                fill_type="solid",
                fgColor=self.COLOR_GREEN_LIGHT,
            )

            cell.font = Font(
                name="Arial",
                size=10,
                bold=True,
                color=self.COLOR_GREEN_TEXT,
            )

        elif normalized_value == "ні":
            cell.fill = PatternFill(
                fill_type="solid",
                fgColor=self.COLOR_RED_LIGHT,
            )

            cell.font = Font(
                name="Arial",
                size=10,
                bold=True,
                color=self.COLOR_RED_TEXT,
            )

    def style_penalty_cell(
        self,
        cell: Any,
    ) -> None:
        """
        Підсвічує штраф більше нуля.
        """

        amount = self.decimal_value(
            cell.value
        )

        if amount <= 0:
            return

        cell.fill = PatternFill(
            fill_type="solid",
            fgColor=self.COLOR_RED_LIGHT,
        )

        cell.font = Font(
            name="Arial",
            size=10,
            bold=True,
            color=self.COLOR_RED_TEXT,
        )

    # ==========================================
    # УМОВНЕ ФОРМАТУВАННЯ
    # ==========================================

    def apply_conditional_formatting(
        self,
        *,
        worksheet: Worksheet,
        headers: tuple[str, ...],
        first_data_row: int,
        last_data_row: int,
    ) -> None:
        """
        Додає Excel conditional formatting.
        """

        for column_index, header in enumerate(
            headers,
            start=1,
        ):
            normalized_header = (
                str(header)
                .strip()
                .lower()
            )

            column_letter = (
                get_column_letter(
                    column_index
                )
            )

            cell_range = (
                f"{column_letter}"
                f"{first_data_row}:"
                f"{column_letter}"
                f"{last_data_row}"
            )

            if (
                "запізнення" in normalized_header
                and "середнє" not in normalized_header
            ):
                worksheet.conditional_formatting.add(
                    cell_range,
                    CellIsRule(
                        operator="greaterThan",
                        formula=["0"],
                        fill=PatternFill(
                            fill_type="solid",
                            fgColor=(
                                self
                                .COLOR_YELLOW_LIGHT
                            ),
                        ),
                        font=Font(
                            color=(
                                self
                                .COLOR_YELLOW_TEXT
                            ),
                            bold=True,
                        ),
                    ),
                )

            if "штраф" in normalized_header:
                worksheet.conditional_formatting.add(
                    cell_range,
                    CellIsRule(
                        operator="greaterThan",
                        formula=["0"],
                        fill=PatternFill(
                            fill_type="solid",
                            fgColor=(
                                self
                                .COLOR_RED_LIGHT
                            ),
                        ),
                        font=Font(
                            color=(
                                self
                                .COLOR_RED_TEXT
                            ),
                            bold=True,
                        ),
                    ),
                )

    # ==========================================
    # ШИРИНА КОЛОНОК
    # ==========================================

    def apply_column_widths(
        self,
        *,
        worksheet: Worksheet,
        widths: tuple[float, ...],
        columns_count: int,
    ) -> None:
        """
        Встановлює читабельну ширину колонок.
        """

        for column_index in range(
            1,
            columns_count + 1,
        ):
            if column_index <= len(widths):
                width = widths[
                    column_index - 1
                ]

            else:
                width = 16

            width = min(
                max(float(width), 8),
                42,
            )

            column_letter = (
                get_column_letter(
                    column_index
                )
            )

            worksheet.column_dimensions[
                column_letter
            ].width = width

    @staticmethod
    def ensure_widths_count(
        *,
        widths: list[float],
        columns_count: int,
    ) -> None:
        """
        Доповнює список ширин колонок.
        """

        while len(widths) < columns_count:
            widths.append(16)

        if len(widths) > columns_count:
            del widths[columns_count:]

    # ==========================================
    # СЛУЖБОВИЙ АРКУШ
    # ==========================================

    def create_metadata_sheet(
        self,
        *,
        workbook: Workbook,
        report_data: ExcelReportData,
    ) -> None:
        """
        Створює прихований службовий аркуш.
        """

        worksheet = workbook.create_sheet(
            title="_metadata"
        )

        rows = [
            (
                "workbook_title",
                report_data.workbook_title,
            ),
            (
                "filename",
                report_data.filename,
            ),
            (
                "penalty_per_minute",
                float(
                    self.PENALTY_PER_MINUTE
                ),
            ),
            (
                "penalty_currency",
                "UAH",
            ),
            (
                "generated_at",
                datetime.now(),
            ),
        ]

        for key, value in (
            report_data.metadata or {}
        ).items():
            rows.append(
                (
                    str(key),
                    self.normalize_excel_value(
                        value
                    ),
                )
            )

        for row_index, (
            key,
            value,
        ) in enumerate(
            rows,
            start=1,
        ):
            worksheet.cell(
                row=row_index,
                column=1,
                value=key,
            )

            worksheet.cell(
                row=row_index,
                column=2,
                value=value,
            )

        worksheet.sheet_state = "hidden"

    # ==========================================
    # НОРМАЛІЗАЦІЯ EXCEL-ЗНАЧЕНЬ
    # ==========================================

    @classmethod
    def normalize_excel_value(
        cls,
        value: Any,
    ) -> Any:
        """
        Перетворює значення у формат openpyxl.
        """

        if value is None:
            return None

        if isinstance(value, Enum):
            return str(value.value)

        if isinstance(value, Decimal):
            return float(value)

        if isinstance(value, Path):
            return str(value)

        if isinstance(
            value,
            (dict, list, tuple, set),
        ):
            return json.dumps(
                value,
                ensure_ascii=False,
                default=str,
            )

        return value

    # ==========================================
    # ФОРМАТУВАННЯ
    # ==========================================

    @staticmethod
    def resolve_alignment(
        normalized_header: str,
    ) -> str:
        """
        Визначає горизонтальне вирівнювання.
        """

        center_markers = (
            "дата",
            "час",
            "статус",
            "тт",
            "id",
            "кластер",
            "кущ",
            "хв",
            "%",
            "каса",
            "штраф",
            "фото",
            "подано",
            "відкриття",
            "закриття",
        )

        if any(
            marker in normalized_header
            for marker in center_markers
        ):
            return "center"

        return "left"

    @staticmethod
    def format_decimal(
        value: Decimal,
    ) -> str:
        """
        Форматує Decimal без зайвих нулів.
        """

        if value == value.to_integral():
            return str(
                int(value)
            )

        return format(
            value,
            "f",
        )

    @staticmethod
    def decimal_value(
        value: Any,
    ) -> Decimal:
        """
        Безпечно перетворює значення в Decimal.
        """

        if value is None:
            return Decimal("0.00")

        try:
            return Decimal(
                str(value)
            ).quantize(
                Decimal("0.01")
            )

        except (
            InvalidOperation,
            TypeError,
            ValueError,
        ):
            return Decimal("0.00")

    # ==========================================
    # НАЗВИ ФАЙЛІВ І АРКУШІВ
    # ==========================================

    @staticmethod
    def sanitize_filename(
        filename: str,
    ) -> str:
        """
        Робить безпечну назву Excel-файлу.
        """

        normalized_filename = (
            filename.strip()
        )

        if not normalized_filename:
            normalized_filename = (
                "soska_bar_report.xlsx"
            )

        normalized_filename = re.sub(
            r'[<>:"/\\|?*]',
            "_",
            normalized_filename,
        )

        if not normalized_filename.lower().endswith(
            ".xlsx"
        ):
            normalized_filename += ".xlsx"

        return normalized_filename[:200]

    @staticmethod
    def unique_sheet_title(
        title: str,
        *,
        used_titles: set[str],
    ) -> str:
        """
        Формує унікальну назву аркуша.
        """

        normalized_title = re.sub(
            r"[\[\]:*?/\\]",
            " ",
            title.strip(),
        )

        normalized_title = " ".join(
            normalized_title.split()
        )

        if not normalized_title:
            normalized_title = "Звіт"

        normalized_title = (
            normalized_title[:31]
        )

        if normalized_title not in used_titles:
            return normalized_title

        base_title = (
            normalized_title[:27]
        )

        counter = 2

        while True:
            candidate = (
                f"{base_title} {counter}"
            )[:31]

            if candidate not in used_titles:
                return candidate

            counter += 1

    # ==========================================
    # ВАЛІДАЦІЯ
    # ==========================================

    @staticmethod
    def validate_report_data(
        report_data: ExcelReportData,
    ) -> None:
        """
        Перевіряє дані звіту.
        """

        if not report_data.workbook_title.strip():
            raise ValueError(
                "Назва Excel-звіту "
                "не може бути порожньою."
            )

        if not report_data.filename.strip():
            raise ValueError(
                "Назва Excel-файлу "
                "не може бути порожньою."
            )

        if not report_data.sheets:
            raise ValueError(
                "Excel-звіт повинен містити "
                "хоча б один аркуш."
            )

        for sheet_data in report_data.sheets:
            if not sheet_data.title.strip():
                raise ValueError(
                    "Назва Excel-аркуша "
                    "не може бути порожньою."
                )

            if not sheet_data.headers:
                raise ValueError(
                    f"Аркуш «{sheet_data.title}» "
                    "не містить заголовків."
                )

            headers_count = len(
                sheet_data.headers
            )

            for row_number, row in enumerate(
                sheet_data.rows,
                start=1,
            ):
                if len(row) != headers_count:
                    raise ValueError(
                        f"Аркуш «{sheet_data.title}», "
                        f"рядок №{row_number}: "
                        "кількість значень не відповідає "
                        "кількості заголовків."
                    )