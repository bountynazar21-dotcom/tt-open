from __future__ import annotations

from copy import copy
from typing import Iterable

from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Protection,
    Side,
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet


# =========================================================
# BASE SIDES / BORDERS
# =========================================================


THIN_SIDE = Side(
    style="thin",
    color="B7B7B7",
)

MEDIUM_SIDE = Side(
    style="medium",
    color="808080",
)

THIN_BORDER = Border(
    left=THIN_SIDE,
    right=THIN_SIDE,
    top=THIN_SIDE,
    bottom=THIN_SIDE,
)

BOTTOM_BORDER = Border(
    bottom=THIN_SIDE,
)

MEDIUM_BOTTOM_BORDER = Border(
    bottom=MEDIUM_SIDE,
)


# =========================================================
# FONTS
# =========================================================


TITLE_FONT = Font(
    name="Arial",
    size=14,
    bold=True,
)

SUBTITLE_FONT = Font(
    name="Arial",
    size=11,
    bold=True,
)

HEADER_FONT = Font(
    name="Arial",
    size=10,
    bold=True,
)

NORMAL_FONT = Font(
    name="Arial",
    size=10,
)

SMALL_FONT = Font(
    name="Arial",
    size=9,
)

BOLD_FONT = Font(
    name="Arial",
    size=10,
    bold=True,
)


# =========================================================
# FILLS
# =========================================================


HEADER_FILL = PatternFill(
    fill_type="solid",
    fgColor="E7E6E6",
)

SUBHEADER_FILL = PatternFill(
    fill_type="solid",
    fgColor="F2F2F2",
)

SUCCESS_FILL = PatternFill(
    fill_type="solid",
    fgColor="E2F0D9",
)

WARNING_FILL = PatternFill(
    fill_type="solid",
    fgColor="FFF2CC",
)

DANGER_FILL = PatternFill(
    fill_type="solid",
    fgColor="F4CCCC",
)

INFO_FILL = PatternFill(
    fill_type="solid",
    fgColor="D9EAF7",
)


# =========================================================
# ALIGNMENT
# =========================================================


ALIGN_LEFT = Alignment(
    horizontal="left",
    vertical="center",
)

ALIGN_CENTER = Alignment(
    horizontal="center",
    vertical="center",
)

ALIGN_RIGHT = Alignment(
    horizontal="right",
    vertical="center",
)

ALIGN_CENTER_WRAP = Alignment(
    horizontal="center",
    vertical="center",
    wrap_text=True,
)

ALIGN_LEFT_WRAP = Alignment(
    horizontal="left",
    vertical="top",
    wrap_text=True,
)


# =========================================================
# PROTECTION
# =========================================================


LOCKED = Protection(
    locked=True,
)

UNLOCKED = Protection(
    locked=False,
)


# =========================================================
# NUMBER FORMATS
# =========================================================


DATE_FORMAT = "DD.MM.YYYY"

TIME_FORMAT = "HH:MM"

DATETIME_FORMAT = "DD.MM.YYYY HH:MM"

MONEY_FORMAT = '#,##0.00'

INTEGER_FORMAT = "0"

PERCENT_FORMAT = "0.00%"


# =========================================================
# CELL HELPERS
# =========================================================


def apply_title_style(
    cell,
) -> None:
    cell.font = copy(
        TITLE_FONT
    )
    cell.alignment = copy(
        ALIGN_LEFT
    )


def apply_subtitle_style(
    cell,
) -> None:
    cell.font = copy(
        SUBTITLE_FONT
    )
    cell.alignment = copy(
        ALIGN_LEFT
    )


def apply_header_style(
    cell,
) -> None:
    cell.font = copy(
        HEADER_FONT
    )
    cell.fill = copy(
        HEADER_FILL
    )
    cell.border = copy(
        THIN_BORDER
    )
    cell.alignment = copy(
        ALIGN_CENTER_WRAP
    )


def apply_normal_style(
    cell,
    *,
    wrap_text: bool = False,
) -> None:
    cell.font = copy(
        NORMAL_FONT
    )
    cell.border = copy(
        THIN_BORDER
    )

    cell.alignment = copy(
        ALIGN_LEFT_WRAP
        if wrap_text
        else ALIGN_LEFT
    )


def apply_center_style(
    cell,
) -> None:
    cell.font = copy(
        NORMAL_FONT
    )
    cell.border = copy(
        THIN_BORDER
    )
    cell.alignment = copy(
        ALIGN_CENTER
    )


def apply_money_style(
    cell,
) -> None:
    apply_center_style(
        cell
    )

    cell.number_format = (
        MONEY_FORMAT
    )


def apply_date_style(
    cell,
) -> None:
    apply_center_style(
        cell
    )

    cell.number_format = (
        DATE_FORMAT
    )


def apply_time_style(
    cell,
) -> None:
    apply_center_style(
        cell
    )

    cell.number_format = (
        TIME_FORMAT
    )


def apply_datetime_style(
    cell,
) -> None:
    apply_center_style(
        cell
    )

    cell.number_format = (
        DATETIME_FORMAT
    )


# =========================================================
# ROW HELPERS
# =========================================================


def style_header_row(
    worksheet: Worksheet,
    row: int = 1,
    *,
    start_column: int = 1,
    end_column: int | None = None,
) -> None:
    """
    Стилізує рядок заголовків.
    """

    if end_column is None:
        end_column = (
            worksheet.max_column
        )

    for column in range(
        start_column,
        end_column + 1,
    ):
        apply_header_style(
            worksheet.cell(
                row=row,
                column=column,
            )
        )


def style_data_rows(
    worksheet: Worksheet,
    *,
    start_row: int = 2,
    end_row: int | None = None,
    start_column: int = 1,
    end_column: int | None = None,
) -> None:
    """
    Додає базовий стиль усім
    клітинкам таблиці.
    """

    if end_row is None:
        end_row = (
            worksheet.max_row
        )

    if end_column is None:
        end_column = (
            worksheet.max_column
        )

    for row in range(
        start_row,
        end_row + 1,
    ):
        for column in range(
            start_column,
            end_column + 1,
        ):
            apply_normal_style(
                worksheet.cell(
                    row=row,
                    column=column,
                )
            )


# =========================================================
# COLUMN WIDTHS
# =========================================================


def set_column_widths(
    worksheet: Worksheet,
    widths: dict[
        int | str,
        float,
    ],
) -> None:
    """
    Встановлює ширину колонок.

    Приклад:

        {
            1: 10,
            2: 25,
            "C": 18,
        }
    """

    for column, width in (
        widths.items()
    ):
        if isinstance(
            column,
            int,
        ):
            letter = (
                get_column_letter(
                    column
                )
            )
        else:
            letter = str(
                column
            ).upper()

        worksheet.column_dimensions[
            letter
        ].width = width


def auto_fit_columns(
    worksheet: Worksheet,
    *,
    min_width: float = 8,
    max_width: float = 50,
    extra_width: float = 2,
) -> None:
    """
    Автоматично підбирає
    ширину колонок.
    """

    for column_cells in (
        worksheet.columns
    ):
        max_length = 0

        first_cell = next(
            iter(
                column_cells
            ),
            None,
        )

        if first_cell is None:
            continue

        column_letter = (
            get_column_letter(
                first_cell.column
            )
        )

        for cell in column_cells:
            value = cell.value

            if value is None:
                continue

            text = str(
                value
            )

            lines = (
                text.splitlines()
                or [text]
            )

            current_length = max(
                len(line)
                for line in lines
            )

            max_length = max(
                max_length,
                current_length,
            )

        width = min(
            max(
                max_length
                + extra_width,
                min_width,
            ),
            max_width,
        )

        worksheet.column_dimensions[
            column_letter
        ].width = width


# =========================================================
# ROW HEIGHTS
# =========================================================


def set_row_height(
    worksheet: Worksheet,
    row: int,
    height: float,
) -> None:
    worksheet.row_dimensions[
        row
    ].height = height


def set_rows_height(
    worksheet: Worksheet,
    rows: Iterable[int],
    height: float,
) -> None:
    for row in rows:
        set_row_height(
            worksheet,
            row,
            height,
        )


# =========================================================
# FREEZE / FILTER
# =========================================================


def freeze_header(
    worksheet: Worksheet,
    *,
    row: int = 1,
) -> None:
    """
    Заморожує рядок заголовків.
    """

    worksheet.freeze_panes = (
        f"A{row + 1}"
    )


def enable_auto_filter(
    worksheet: Worksheet,
    *,
    start_row: int = 1,
) -> None:
    """
    Вмикає Excel-фільтр
    для таблиці.
    """

    if (
        worksheet.max_row
        < start_row
        or worksheet.max_column
        < 1
    ):
        return

    end_column = (
        get_column_letter(
            worksheet.max_column
        )
    )

    worksheet.auto_filter.ref = (
        f"A{start_row}:"
        f"{end_column}"
        f"{worksheet.max_row}"
    )


# =========================================================
# STATUS COLORS
# =========================================================


def apply_success_fill(
    cell,
) -> None:
    cell.fill = copy(
        SUCCESS_FILL
    )


def apply_warning_fill(
    cell,
) -> None:
    cell.fill = copy(
        WARNING_FILL
    )


def apply_danger_fill(
    cell,
) -> None:
    cell.fill = copy(
        DANGER_FILL
    )


def apply_info_fill(
    cell,
) -> None:
    cell.fill = copy(
        INFO_FILL
    )


def apply_status_fill(
    cell,
    status: object,
) -> None:
    """
    Базова підсвітка статусів.

    Працює і з Enum, і зі string.
    """

    value = getattr(
        status,
        "value",
        status,
    )

    normalized = str(
        value or ""
    ).strip().upper()

    if any(
        item in normalized
        for item in (
            "ON_TIME",
            "OPENED_EARLY",
            "SUBMITTED_ON_TIME",
            "SUCCESS",
            "ACTIVE",
            "CONFIRMED",
        )
    ):
        apply_success_fill(
            cell
        )

    elif any(
        item in normalized
        for item in (
            "LATE",
            "WARNING",
            "WAITING",
            "PENDING",
        )
    ):
        apply_warning_fill(
            cell
        )

    elif any(
        item in normalized
        for item in (
            "MISSED",
            "ERROR",
            "FAILED",
            "BLOCKED",
            "REJECTED",
        )
    ):
        apply_danger_fill(
            cell
        )

    else:
        apply_info_fill(
            cell
        )


# =========================================================
# SHEET PRESET
# =========================================================


def prepare_report_sheet(
    worksheet: Worksheet,
    *,
    header_row: int = 1,
    freeze: bool = True,
    auto_filter: bool = True,
    auto_width: bool = True,
) -> Worksheet:
    """
    Базова підготовка Excel-листа
    для звіту.
    """

    style_header_row(
        worksheet,
        row=header_row,
    )

    if (
        worksheet.max_row
        > header_row
    ):
        style_data_rows(
            worksheet,
            start_row=(
                header_row + 1
            ),
        )

    if freeze:
        freeze_header(
            worksheet,
            row=header_row,
        )

    if auto_filter:
        enable_auto_filter(
            worksheet,
            start_row=header_row,
        )

    if auto_width:
        auto_fit_columns(
            worksheet
        )

    return worksheet


# =========================================================
# EXPORTS
# =========================================================


__all__ = [
    "THIN_SIDE",
    "MEDIUM_SIDE",
    "THIN_BORDER",
    "BOTTOM_BORDER",
    "MEDIUM_BOTTOM_BORDER",

    "TITLE_FONT",
    "SUBTITLE_FONT",
    "HEADER_FONT",
    "NORMAL_FONT",
    "SMALL_FONT",
    "BOLD_FONT",

    "HEADER_FILL",
    "SUBHEADER_FILL",
    "SUCCESS_FILL",
    "WARNING_FILL",
    "DANGER_FILL",
    "INFO_FILL",

    "ALIGN_LEFT",
    "ALIGN_CENTER",
    "ALIGN_RIGHT",
    "ALIGN_CENTER_WRAP",
    "ALIGN_LEFT_WRAP",

    "LOCKED",
    "UNLOCKED",

    "DATE_FORMAT",
    "TIME_FORMAT",
    "DATETIME_FORMAT",
    "MONEY_FORMAT",
    "INTEGER_FORMAT",
    "PERCENT_FORMAT",

    "apply_title_style",
    "apply_subtitle_style",
    "apply_header_style",
    "apply_normal_style",
    "apply_center_style",
    "apply_money_style",
    "apply_date_style",
    "apply_time_style",
    "apply_datetime_style",

    "style_header_row",
    "style_data_rows",

    "set_column_widths",
    "auto_fit_columns",

    "set_row_height",
    "set_rows_height",

    "freeze_header",
    "enable_auto_filter",

    "apply_success_fill",
    "apply_warning_fill",
    "apply_danger_fill",
    "apply_info_fill",
    "apply_status_fill",

    "prepare_report_sheet",
]